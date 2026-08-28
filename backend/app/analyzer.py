from __future__ import annotations

import csv
import io
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MAX_SAMPLE_ROWS = 100


def _slug(value: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9]+", "_", value.strip()).strip("_")
    return value.lower() or "field"


def _python_type(value: Any) -> str:
    if value is None or value == "":
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int) and not isinstance(value, bool):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, (datetime, date)):
        return "datetime"
    return "string"


def _infer_type(values: list[Any]) -> str:
    types = [_python_type(v) for v in values if v not in (None, "")]
    if not types:
        return "string"
    counts = Counter(types)
    if set(counts).issubset({"integer", "number"}):
        return "number" if "number" in counts else "integer"
    return counts.most_common(1)[0][0]


def _guess_primary_key(headers: list[str], rows: list[list[Any]]) -> str | None:
    if not headers:
        return None
    candidates = []
    for index, header in enumerate(headers):
        normalized = _slug(header)
        score = 0
        if normalized == "id":
            score += 5
        if normalized.startswith("id_") or normalized.endswith("_id"):
            score += 4
        values = [row[index] for row in rows if index < len(row) and row[index] not in (None, "")]
        if values and len(values) == len(set(map(str, values))):
            score += 3
        if values and len(values) >= max(1, len(rows) // 2):
            score += 1
        candidates.append((score, normalized))
    best = max(candidates, default=(0, None))
    return best[1] if best[0] >= 4 else None


def _sheet_to_entity(name: str, rows: list[list[Any]], formula_count: int = 0) -> dict[str, Any]:
    if not rows:
        return {"name": _slug(name), "source_name": name, "fields": [], "primary_key": None, "row_count": 0, "formula_count": formula_count}

    headers = [str(value).strip() if value not in (None, "") else f"column_{i + 1}" for i, value in enumerate(rows[0])]
    data_rows = rows[1 : MAX_SAMPLE_ROWS + 1]
    fields = []

    for index, header in enumerate(headers):
        sample = [row[index] if index < len(row) else None for row in data_rows]
        fields.append(
            {
                "name": _slug(header),
                "source_name": header,
                "type": _infer_type(sample),
                "nullable": any(value in (None, "") for value in sample) if sample else True,
                "sample_values": [value for value in sample if value not in (None, "")][:3],
            }
        )

    return {
        "name": _slug(name),
        "source_name": name,
        "fields": fields,
        "primary_key": _guess_primary_key(headers, data_rows),
        "row_count": max(0, len(rows) - 1),
        "formula_count": formula_count,
    }


def _detect_relationships(entities: list[dict[str, Any]]) -> list[dict[str, str]]:
    relationships: list[dict[str, str]] = []
    entity_names = {entity["name"]: entity for entity in entities}

    for entity in entities:
        for field in entity["fields"]:
            field_name = field["name"]
            if not field_name.endswith("_id"):
                continue
            target_name = field_name[:-3]
            target = entity_names.get(target_name)
            if not target:
                target = next((item for name, item in entity_names.items() if name.rstrip("s") == target_name.rstrip("s")), None)
            if target:
                relationships.append(
                    {
                        "from_entity": entity["name"],
                        "from_field": field_name,
                        "to_entity": target["name"],
                        "to_field": target.get("primary_key") or "id",
                        "kind": "many-to-one",
                    }
                )
    return relationships


def analyze_xlsx(content: bytes, filename: str) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=False)
    entities = []

    for worksheet in workbook.worksheets:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        while rows and not any(value not in (None, "") for value in rows[0]):
            rows.pop(0)
        formula_count = sum(
            1
            for row in worksheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        entities.append(_sheet_to_entity(worksheet.title, rows, formula_count))

    return {
        "workbook": {"filename": filename, "type": "xlsx", "sheet_count": len(workbook.sheetnames), "sheets": workbook.sheetnames},
        "entities": entities,
        "relationships": _detect_relationships(entities),
    }


def analyze_csv(content: bytes, filename: str) -> dict[str, Any]:
    text = content.decode("utf-8-sig")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    entity = _sheet_to_entity(Path(filename).stem, rows)
    return {
        "workbook": {"filename": filename, "type": "csv", "sheet_count": 1, "sheets": [Path(filename).stem]},
        "entities": [entity],
        "relationships": [],
    }


def analyze_file(content: bytes, filename: str) -> dict[str, Any]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        return analyze_xlsx(content, filename)
    if suffix == ".csv":
        return analyze_csv(content, filename)
    raise ValueError("Formato não suportado. Use .xlsx ou .csv.")
