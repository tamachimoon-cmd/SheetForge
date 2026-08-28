from __future__ import annotations

import csv
import io
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def _serialize_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def _trim_leading_empty_rows(rows: list[list[Any]]) -> list[list[Any]]:
    while rows and not any(value not in (None, "") for value in rows[0]):
        rows.pop(0)
    return rows


def _rows_to_records(rows: list[list[Any]], fields: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = _trim_leading_empty_rows(rows)
    if not rows:
        return []

    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue
        record = {
            field["name"]: _serialize_value(row[index] if index < len(row) else None)
            for index, field in enumerate(fields)
        }
        records.append(record)
    return records


def _read_xlsx(content: bytes, app_schema: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    result: dict[str, list[dict[str, Any]]] = {}

    for entity in app_schema["entities"]:
        source_name = entity["label"]
        if source_name not in workbook.sheetnames:
            result[entity["name"]] = []
            continue
        worksheet = workbook[source_name]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        result[entity["name"]] = _rows_to_records(rows, entity["fields"])

    return result


def _read_csv(content: bytes, app_schema: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    rows = [list(row) for row in csv.reader(io.StringIO(text), dialect)]
    entity = app_schema["entities"][0]
    return {entity["name"]: _rows_to_records(rows, entity["fields"])}


def read_source_rows(content: bytes, filename: str, app_schema: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(content, app_schema)
    if suffix == ".csv":
        return _read_csv(content, app_schema)
    raise ValueError("Formato não suportado para importação. Use .xlsx ou .csv.")
